import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io

# --- Настройка страницы ---
st.set_page_config(page_title="Таможенный Заполнитель Pro", page_icon="🗃️", layout="wide")

st.title("📦 Автоматическая подготовка данных (с выбором листов Excel)")
st.markdown("Загрузите исходные документы. Если в файлах несколько листов, вы сможете выбрать нужный прямо в интерфейсе.")

# --- Вспомогательная функция для надежного получения списка листов ---
def get_excel_sheets(uploaded_file):
    uploaded_file.seek(0)
    file_name = uploaded_file.name.lower()
    # Жестко определяем движок по расширению
    engine_type = 'xlrd' if file_name.endswith('.xls') else 'openpyxl'
    try:
        xl = pd.ExcelFile(uploaded_file, engine=engine_type)
        return xl.sheet_names
    except ValueError:
        # Запасной план на случай обманчивого расширения (например, .xls.xlsx)
        uploaded_file.seek(0)
        fallback = 'openpyxl' if engine_type == 'xlrd' else 'xlrd'
        xl = pd.ExcelFile(uploaded_file, engine=fallback)
        return xl.sheet_names

# --- Безопасный поиск колонок ---
def find_col(columns, keywords):
    for col in columns:
        col_clean = str(col).lower().strip()
        if any(kw in col_clean for kw in keywords):
            return col
    return None

# --- Парсинг выбранного листа со смещенной шапкой (С ИСПРАВЛЕНИЕМ ОШИБКИ ДВИЖКА) ---
def parse_selected_sheet(uploaded_file, sheet_name=None):
    uploaded_file.seek(0)
    file_name = uploaded_file.name.lower()
    
    if file_name.endswith('.csv'):
        try:
            df = pd.read_csv(uploaded_file, header=None, sep=None, engine='python')
        except:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, header=None)
    else:
        # Явное указание engine спасает от ошибки "Excel file format cannot be determined"
        engine_type = 'xlrd' if file_name.endswith('.xls') else 'openpyxl'
        try:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, engine=engine_type)
        except ValueError:
            uploaded_file.seek(0)
            fallback = 'openpyxl' if engine_type == 'xlrd' else 'xlrd'
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, engine=fallback)
        
    header_row_idx = 0
    for idx, row in df.iterrows():
        row_str = " ".join(row.astype(str)).lower()
        if any(kw in row_str for kw in ['код изделия', 'part number', 'part no', 'наименование', 'goods name', '№ места', 'package no']):
            header_row_idx = idx
            break
            
    df_clean = df.iloc[header_row_idx:].copy()
    df_clean.columns = df_clean.iloc[0]
    df_clean = df_clean.iloc[1:].reset_index(drop=True)
    return df_clean

# --- Безопасное приведение к числу (защита от запятых вместо точек) ---
def to_numeric_safe(series):
    s = series.astype(str).str.replace(',', '.').str.replace(' ', '')
    return pd.to_numeric(s, errors='coerce')

# --- Функция генерации Excel ---
def create_styled_excel(df_data, box_summary_data):
    wb = openpyxl.Workbook()
    
    # Стили
    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F497D")
    SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True)
    REGULAR_FONT = Font(name="Calibri", size=11)
    ZEBRA_FILL = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    border_all = Border(left=Side(border_style="thin", color="D9D9D9"), right=Side(border_style="thin", color="D9D9D9"), 
                        top=Side(border_style="thin", color="D9D9D9"), bottom=Side(border_style="thin", color="D9D9D9"))
    border_total = Border(top=Side(border_style="thin", color="1F497D"), bottom=Side(border_style="double", color="1F497D"), 
                          left=Side(border_style="thin", color="D9D9D9"), right=Side(border_style="thin", color="D9D9D9"))
    
    # === ЛИСТ 1 ===
    ws_data = wb.active
    ws_data.title = "Данные для Заполнителя"
    ws_data.views.sheetView[0].showGridLines = True
    
    ws_data["A1"] = "Таблица данных для программы 'Заполнитель' (Альта-Софт)"
    ws_data["A1"].font = TITLE_FONT
    ws_data["A2"] = "Сформировано автоматически на основе загруженных документов"
    ws_data["A2"].font = SUBTITLE_FONT
    
    headers = [
        "Порядковый номер", "Наименование товара", "Числовой код ед. изм.", 
        "Буквенный код ед. изм.", "Количество товара", "Цена за единицу, USD", 
        "Общая стоимость, USD", "Код изделия / артикул", "Вес нетто за ед., кг", 
        "Общий вес нетто, кг", "Общий вес брутто, кг", "Номер грузового места"
    ]
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws_data.cell(row=4, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_data.row_dimensions[4].height = 40
    
    start_row = 5
    for i, row in df_data.iterrows():
        current_row = start_row + i
        
        ws_data.cell(row=current_row, column=1, value=i+1).alignment = Alignment(horizontal="center")
        ws_data.cell(row=current_row, column=2, value=str(row["name"])).alignment = Alignment(horizontal="left")
        ws_data.cell(row=current_row, column=3, value=row["code_num"]).alignment = Alignment(horizontal="center")
        ws_data.cell(row=current_row, column=4, value=row["code_str"]).alignment = Alignment(horizontal="center")
        ws_data.cell(row=current_row, column=5, value=row["qty"]).alignment = Alignment(horizontal="right")
        ws_data.cell(row=current_row, column=6, value=row["price"]).alignment = Alignment(horizontal="right")
        ws_data.cell(row=current_row, column=7, value=f"=E{current_row}*F{current_row}").alignment = Alignment(horizontal="right")
        ws_data.cell(row=current_row, column=8, value=str(row["part_no"])).alignment = Alignment(horizontal="center")
        ws_data.cell(row=current_row, column=9, value=row["net_unit"]).alignment = Alignment(horizontal="right")
        ws_data.cell(row=current_row, column=10, value=f"=E{current_row}*I{current_row}").alignment = Alignment(horizontal="right")
        ws_data.cell(row=current_row, column=11, value=row["gross_total"]).alignment = Alignment(horizontal="right")
        ws_data.cell(row=current_row, column=12, value=str(row["box_num"])).alignment = Alignment(horizontal="center")
        
        for col_idx in range(1, 13):
            cell = ws_data.cell(row=current_row, column=col_idx)
            cell.font = REGULAR_FONT
            cell.border = border_all
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
                
        ws_data.cell(row=current_row, column=5).number_format = '#,##0'
        ws_data.cell(row=current_row, column=6).number_format = '#,##0.00'
        ws_data.cell(row=current_row, column=7).number_format = '#,##0.00'
        ws_data.cell(row=current_row, column=9).number_format = '#,##0.00'
        ws_data.cell(row=current_row, column=10).number_format = '#,##0.00'
        ws_data.cell(row=current_row, column=11).number_format = '#,##0.000'
        ws_data.cell(row=current_row, column=3).number_format = '@'
        ws_data.cell(row=current_row, column=8).number_format = '@'
        
    tot_row = start_row + len(df_data)
    ws_data.cell(row=tot_row, column=2, value="ИТОГО:").font = BOLD_FONT
    ws_data.cell(row=tot_row, column=2).alignment = Alignment(horizontal="right")
    ws_data.cell(row=tot_row, column=5, value=f"=SUM(E5:E{tot_row-1})").number_format = '#,##0'
    ws_data.cell(row=tot_row, column=7, value=f"=SUM(G5:G{tot_row-1})").number_format = '#,##0.00'
    ws_data.cell(row=tot_row, column=10, value=f"=SUM(J5:J{tot_row-1})").number_format = '#,##0.00'
    ws_data.cell(row=tot_row, column=11, value=f"=SUM(K5:K{tot_row-1})").number_format = '#,##0.000'
    
    for col_idx in range(1, 13):
        cell = ws_data.cell(row=tot_row, column=col_idx)
        cell.font = BOLD_FONT
        cell.border = border_total
        cell.fill = TOTAL_FILL
        if col_idx in [5, 7, 10, 11]:
            cell.alignment = Alignment(horizontal="right")
            
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row > 2)
        ws_data.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)
    ws_data.column_dimensions['B'].width = 45

    # === ЛИСТ 2 ===
    ws_dash = wb.create_sheet(title="Сводная аналитика")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash["A1"] = "Сводный отчет по товарной партии"
    ws_dash["A1"].font = TITLE_FONT
    
    kpis = [
        ("Общая стоимость партии", f"='Данные для Заполнителя'!G{tot_row}", "#,##0.00 USD", "B3", "B4"),
        ("Общий вес нетто", f"='Данные для Заполнителя'!J{tot_row}", "#,##0.00 кг", "D3", "D4"),
        ("Общий вес брутто", f"='Данные для Заполнителя'!K{tot_row}", "#,##0.00 кг", "F3", "F4")
    ]
    
    for label, formula, num_fmt, l_cell, v_cell in kpis:
        ws_dash[l_cell] = label
        ws_dash[l_cell].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws_dash[l_cell].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws_dash[l_cell].alignment = Alignment(horizontal="center", vertical="center")
        
        ws_dash[v_cell] = formula
        ws_dash[v_cell].font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        ws_dash[v_cell].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        ws_dash[v_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash[v_cell].number_format = num_fmt
        ws_dash[v_cell].border = border_all
        
    ws_dash.row_dimensions[3].height = 20
    ws_dash.row_dimensions[4].height = 30
    
    ws_dash["B7"] = "Сводка по грузовым местам"
    ws_dash["B7"].font = BOLD_FONT
    
    dash_headers = ["Грузовое место", "Кол-во позиций", "Вес нетто, кг", "Вес брутто, кг"]
    for col_idx, text in enumerate(dash_headers, 2):
        cell = ws_dash.cell(row=8, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        
    d_idx = 0
    for box, info in box_summary_data.items():
        r = 9 + d_idx
        ws_dash.cell(row=r, column=2, value=box).alignment = Alignment(horizontal="center")
        ws_dash.cell(row=r, column=3, value=info["items_count"]).alignment = Alignment(horizontal="right")
        ws_dash.cell(row=r, column=4, value=info["net_total"]).alignment = Alignment(horizontal="right")
        ws_dash.cell(row=r, column=5, value=info["gross_total"]).alignment = Alignment(horizontal="right")
        
        for c in range(2, 6):
            cell = ws_dash.cell(row=r, column=c)
            cell.font = REGULAR_FONT
            cell.border = border_all
            if c >= 4: cell.number_format = '#,##0.00'
        d_idx += 1
        
    r_tot = 9 + d_idx
    ws_dash.cell(row=r_tot, column=2, value="Всего").font = BOLD_FONT
    ws_dash.cell(row=r_tot, column=2).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=r_tot, column=3, value=f"=SUM(C9:C{r_tot-1})")
    ws_dash.cell(row=r_tot, column=4, value=f"=SUM(D9:D{r_tot-1})")
    ws_dash.cell(row=r_tot, column=5, value=f"=SUM(E9:E{r_tot-1})")
    
    for c in range(2, 6):
        cell = ws_dash.cell(row=r_tot, column=c)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = border_total
        if c >= 3:
            cell.alignment = Alignment(horizontal="right")
            if c >= 4: cell.number_format = '#,##0.00'
            
    if len(box_summary_data) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Распределение веса по грузовым местам (кг)"
        chart.y_axis.title = "Вес, кг"
        chart.x_axis.title = "Грузовое место"
        
        data = Reference(ws_dash, min_col=4, min_row=8, max_col=5, max_row=8+len(box_summary_data))
        cats = Reference(ws_dash, min_col=2, min_row=9, max_row=8+len(box_summary_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_dash.add_chart(chart, "B14")
        chart.width = 16
        chart.height = 10
            
    for col in ws_dash.columns:
        ws_dash.column_dimensions[get_column_letter(col[0].column)].width = 18
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- Интерфейс приложения ---
col1, col2, col3 = st.columns(3)

spec_sheet, invoice_sheet, pack_sheet = None, None, None

with col1:
    spec_file = st.file_uploader("📋 Спецификация", type=["xlsx", "xls", "csv"])
    if spec_file and not spec_file.name.lower().endswith('.csv'):
        sheets = get_excel_sheets(spec_file)
        if len(sheets) > 1:
            spec_sheet = st.selectbox("Выберите лист Спецификации:", sheets, key="spec_s")
        else:
            spec_sheet = sheets[0]

with col2:
    invoice_file = st.file_uploader("🧾 Инвойс", type=["xlsx", "xls", "csv"])
    if invoice_file and not invoice_file.name.lower().endswith('.csv'):
        sheets = get_excel_sheets(invoice_file)
        if len(sheets) > 1:
            invoice_sheet = st.selectbox("Выберите лист Инвойса:", sheets, key="inv_s")
        else:
            invoice_sheet = sheets[0]

with col3:
    pack_file = st.file_uploader("📦 Упаковочный лист", type=["xlsx", "xls", "csv"])
    if pack_file and not pack_file.name.lower().endswith('.csv'):
        sheets = get_excel_sheets(pack_file)
        if len(sheets) > 1:
            pack_sheet = st.selectbox("Выберите лист Упаковочного:", sheets, key="pack_s")
        else:
            pack_sheet = sheets[0]

if spec_file and invoice_file and pack_file:
    if st.button("Сформировать отчет", type="primary"):
        with st.spinner('Анализируем данные с выбранных листов...'):
            try:
                # 1. Читаем выбранный лист Упаковочного
                df_p = parse_selected_sheet(pack_file, sheet_name=pack_sheet)
                
                c_box = find_col(df_p.columns, ['мест', 'package', 'box'])
                c_part = find_col(df_p.columns, ['код изделия', 'part number', 'part no', 'артикул'])
                c_name = find_col(df_p.columns, ['наименование', 'goods name', 'description'])
                c_qty = find_col(df_p.columns, ['кол', 'qty', 'quantity'])
                c_net = find_col(df_p.columns, ['нетто', 'net weight', 'net'])
                c_gross = find_col(df_p.columns, ['брутто', 'gross weight', 'gross'])
                
                required_cols = {
                    "Номер места": c_box, "Артикул": c_part, "Наименование": c_name, 
                    "Количество": c_qty, "Вес нетто": c_net, "Вес брутто": c_gross
                }
                
                missing_cols = [name for name, val in required_cols.items() if val is None]
                if missing_cols:
                    st.error(f"⚠️ На выбранном листе не найдена колонка: **{', '.join(missing_cols)}** в Упаковочном. Проверьте правильность выбора листа.")
                    st.stop()
                    
                df_p[c_box] = df_p[c_box].ffill()
                df_p[c_qty] = to_numeric_safe(df_p[c_qty])
                df_p[c_net] = to_numeric_safe(df_p[c_net])
                df_p[c_gross] = to_numeric_safe(df_p[c_gross])
                
                df_p = df_p.dropna(subset=[c_part, c_qty])
                
                box_gross_weights = df_p.groupby(c_box)[c_gross].max().to_dict()
                df_p_grouped = df_p.groupby([c_box, c_part, c_name], as_index=False).agg({c_qty: 'sum', c_net: 'sum'})

                # 2. Собираем цены из Спецификации и Инвойса
                price_dict = {}
                docs_config = [(spec_file, spec_sheet), (invoice_file, invoice_sheet)]
                
                for doc, sheet in docs_config:
                    try:
                        df_doc = parse_selected_sheet(doc, sheet_name=sheet)
                        c_doc_part = find_col(df_doc.columns, ['код', 'part', 'артикул'])
                        c_doc_price = find_col(df_doc.columns, ['цена', 'price', 'тариф'])
                        
                        if c_doc_part and c_doc_price:
                            df_doc[c_doc_price] = to_numeric_safe(df_doc[c_doc_price])
                            extracted_prices = df_doc.dropna(subset=[c_doc_part, c_doc_price]).set_index(c_doc_part)[c_doc_price].to_dict()
                            price_dict.update(extracted_prices)
                    except:
                        continue

                # 3. Основной цикл расчетов
                final_rows = []
                box_summary = {}
                unique_boxes = df_p_grouped[c_box].unique()
                
                for box in unique_boxes:
                    box_df = df_p_grouped[df_p_grouped[c_box] == box]
                    box_gross = float(box_gross_weights.get(box, 0.0) or 0.0)
                    box_net_total = float(box_df[c_net].sum() or 0.0)
                    
                    indices = box_df.index.tolist()
                    running_gross = 0.0
                    
                    for idx in indices[:-1]:
                        row = box_df.loc[idx]
                        item_net = float(row[c_net]) if pd.notnull(row[c_net]) else 0.0
                        
                        calc_gross = round(item_net * (box_gross / box_net_total), 3) if box_net_total > 0 else 0.0
                        running_gross += calc_gross
                        
                        part_val = str(row[c_part]).strip()
                        qty_val = int(row[c_qty]) if pd.notnull(row[c_qty]) else 0
                        price_val = float(price_dict.get(part_val, 0.0))
                        
                        final_rows.append({
                            "name": str(row[c_name]), "code_num": "796", "code_str": "шт",
                            "qty": qty_val, "price": price_val, "part_no": part_val,
                            "net_unit": round(item_net / qty_val, 3) if qty_val > 0 else 0.0,
                            "box_num": str(box), "gross_total": calc_gross
                        })
                        
                    if indices:
                        last_idx = indices[-1]
                        row = box_df.loc[last_idx]
                        item_net = float(row[c_net]) if pd.notnull(row[c_net]) else 0.0
                        calc_gross = round(box_gross - running_gross, 3)
                        
                        part_val = str(row[c_part]).strip()
                        qty_val = int(row[c_qty]) if pd.notnull(row[c_qty]) else 0
                        price_val = float(price_dict.get(part_val, 0.0))
                        
                        final_rows.append({
                            "name": str(row[c_name]), "code_num": "796", "code_str": "шт",
                            "qty": qty_val, "price": price_val, "part_no": part_val,
                            "net_unit": round(item_net / qty_val, 3) if qty_val > 0 else 0.0,
                            "box_num": str(box), "gross_total": calc_gross
                        })
                        
                    box_summary[f"Место {box}"] = {
                        "items_count": len(indices), "net_total": box_net_total, "gross_total": box_gross
                    }

                df_final_result = pd.DataFrame(final_rows)
                
                if df_final_result.empty:
                    st.error("⚠️ Не удалось сопоставить данные. Убедитесь, что артикулы в файлах совпадают.")
                    st.stop()

                # 4. Выгрузка
                excel_output = create_styled_excel(df_final_result, box_summary)
                
                st.success("🎉 Успех! Файл сформирован с учетом всех данных.")
                st.download_button(
                    label="📥 Скачать готовый Excel",
                    data=excel_output,
                    file_name="Tamozhnya_MultiSheet_Result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"Системная ошибка: {e}")
                st.info("Пожалуйста, убедитесь, что в файле requirements.txt прописана библиотека xlrd")
